"""
Bank Statement Parser
Supports parsing bank statements from multiple banks in various formats (PDF, Excel, CSV)
"""
from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional
from datetime import datetime
import re
import logging
import pandas as pd
import PyPDF2
from io import BytesIO
from app.models.expense import ExpenseType, PaymentMethod

logger = logging.getLogger(__name__)


class BankStatementParser(ABC):
    """Base class for bank statement parsers"""

    def __init__(self, file_path: str, password: str = None):
        self.file_path = file_path
        self.password = password
        self.transactions: List[Dict[str, Any]] = []
    
    @abstractmethod
    def parse(self) -> List[Dict[str, Any]]:
        """Parse the bank statement and return list of transactions"""
        pass
    
    def _clean_amount(self, amount_str: str) -> float:
        """Clean and convert amount string to float"""
        if not amount_str or amount_str.strip() == '':
            return 0.0
        
        # Remove currency symbols, commas, and spaces
        cleaned = re.sub(r'[₹$,\s]', '', str(amount_str))
        
        # Handle negative amounts in parentheses
        if '(' in cleaned and ')' in cleaned:
            cleaned = '-' + cleaned.replace('(', '').replace(')', '')
        
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    
    def _parse_date(self, date_str: str, formats: List[str]) -> Optional[datetime]:
        """Try to parse date with multiple formats"""
        for fmt in formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue
        return None
    
    def _detect_transaction_type(self, description: str, debit: float, credit: float) -> ExpenseType:
        """Detect transaction type based on description and amounts"""
        description_lower = description.lower()
        
        # Check for transfers
        transfer_keywords = ['transfer', 'neft', 'imps', 'rtgs', 'upi transfer']
        if any(keyword in description_lower for keyword in transfer_keywords):
            return ExpenseType.TRANSFER
        
        # Check amounts
        if credit > 0 and debit == 0:
            return ExpenseType.CREDIT
        elif debit > 0 and credit == 0:
            return ExpenseType.DEBIT
        
        return ExpenseType.DEBIT
    
    def _detect_payment_method(self, description: str) -> Optional[PaymentMethod]:
        """Detect payment method from transaction description"""
        description_lower = description.lower()
        
        if 'upi' in description_lower or 'paytm' in description_lower or 'phonepe' in description_lower or 'gpay' in description_lower:
            return PaymentMethod.UPI
        elif 'neft' in description_lower or 'imps' in description_lower or 'rtgs' in description_lower:
            return PaymentMethod.NET_BANKING
        elif 'atm' in description_lower or 'pos' in description_lower or 'card' in description_lower:
            return PaymentMethod.CREDIT_CARD
        elif 'cheque' in description_lower or 'chq' in description_lower:
            return PaymentMethod.CHEQUE
        
        return None
    
    def _extract_merchant_name(self, description: str) -> Optional[str]:
        """Extract merchant name from transaction description"""
        # Remove common prefixes
        prefixes = ['upi-', 'pos-', 'atm-', 'neft-', 'imps-', 'rtgs-']
        desc_lower = description.lower()
        
        for prefix in prefixes:
            if desc_lower.startswith(prefix):
                description = description[len(prefix):]
                break
        
        # Extract merchant name (usually before @ or first few words)
        if '@' in description:
            merchant = description.split('@')[0].strip()
        else:
            # Take first 3-5 words as merchant name
            words = description.split()[:5]
            merchant = ' '.join(words)
        
        return merchant.strip() if merchant else None


class ICICIBankParser(BankStatementParser):
    """Parser for ICICI Bank statements (PDF and Excel)"""
    
    def parse(self) -> List[Dict[str, Any]]:
        """Parse ICICI Bank statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        elif self.file_path.lower().endswith(('.xlsx', '.xls')):
            return self._parse_excel()
        else:
            raise ValueError("Unsupported file format for ICICI Bank")
    
    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse ICICI Bank PDF statement — tries table extraction first, falls back to text"""
        transactions = self._parse_pdf_tables()
        if transactions:
            logger.info(f"ICICI Bank PDF: table extraction found {len(transactions)} transactions")
            return transactions

        transactions = self._parse_pdf_text()
        logger.info(f"ICICI Bank PDF: text extraction found {len(transactions)} transactions")
        return transactions

    def _parse_pdf_tables(self) -> List[Dict[str, Any]]:
        """Parse ICICI Bank PDF using pdfplumber table extraction"""
        import pdfplumber
        transactions = []

        try:
            with pdfplumber.open(self.file_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        for row in table:
                            if not row or len(row) < 7:
                                continue
                            # Columns: S No. | Value Date | Transaction Date | Cheque Number |
                            #          Transaction Remarks | Withdrawal Amount(INR) | Deposit Amount(INR) | Balance(INR)
                            # Check if first cell is a serial number or if we have dates in expected positions
                            cells = [str(c).strip() if c else '' for c in row]

                            # Find the row layout — may or may not have S No. column
                            value_date_str = ''
                            txn_date_str = ''
                            description = ''
                            withdrawal_str = ''
                            deposit_str = ''
                            balance_str = ''

                            # Try 8-column layout (with S No.)
                            if len(cells) >= 8 and re.match(r'\d{2}/\d{2}/\d{4}', cells[1]):
                                value_date_str = cells[1]
                                txn_date_str = cells[2]
                                description = cells[4]
                                withdrawal_str = cells[5]
                                deposit_str = cells[6]
                                balance_str = cells[7]
                            # Try 7-column layout (without S No.)
                            elif len(cells) >= 7 and re.match(r'\d{2}/\d{2}/\d{4}', cells[0]):
                                value_date_str = cells[0]
                                txn_date_str = cells[1]
                                description = cells[3]
                                withdrawal_str = cells[4]
                                deposit_str = cells[5]
                                balance_str = cells[6]
                            else:
                                continue

                            date_str = txn_date_str if txn_date_str else value_date_str
                            if not re.match(r'\d{2}/\d{2}/\d{4}', date_str):
                                continue

                            # Clean description (replace newlines from multi-line cells)
                            description = re.sub(r'\s*\n\s*', ' ', description).strip()
                            if not description:
                                continue

                            debit = self._clean_amount(withdrawal_str) if withdrawal_str and withdrawal_str not in ('', 'None', 'nan') else 0.0
                            credit = self._clean_amount(deposit_str) if deposit_str and deposit_str not in ('', 'None', 'nan') else 0.0
                            balance = self._clean_amount(balance_str) if balance_str and balance_str not in ('', 'None', 'nan') else 0.0

                            transaction_date = self._parse_date(date_str, ['%d/%m/%Y'])
                            if transaction_date and (debit > 0 or credit > 0):
                                transactions.append({
                                    'transaction_date': transaction_date,
                                    'description': description,
                                    'amount': debit if debit > 0 else credit,
                                    'transaction_type': self._detect_transaction_type(description, debit, credit),
                                    'payment_method': self._detect_payment_method(description),
                                    'merchant_name': self._extract_merchant_name(description),
                                    'reference_number': None,
                                    'balance_after': balance
                                })
        except Exception as e:
            logger.error(f"ICICI Bank PDF table extraction failed: {e}")
            return []

        return transactions

    def _parse_pdf_text(self) -> List[Dict[str, Any]]:
        """Parse ICICI Bank PDF using text extraction with multi-line description handling"""
        import pdfplumber
        transactions = []

        # Collect all lines from all pages
        all_lines = []
        try:
            with pdfplumber.open(self.file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        all_lines.extend(text.split('\n'))
        except Exception as e:
            logger.error(f"ICICI Bank PDF text extraction failed: {e}")
            return []

        # Transaction line pattern:
        # S.No  ValueDate  TxnDate  ChequeNo  Description  Withdrawal  Deposit  Balance
        # The S.No may or may not be present. Dates are DD/MM/YYYY.
        # Sometimes: 1 01/12/2025 01/12/2025 UPI/RATIKANTA/... 0.00 16851.00 199956.35
        # Key: line starts with optional S.No, then two dates (value date + txn date)
        txn_re = re.compile(
            r'(?:\d+\s+)?'                         # optional S No.
            r'(\d{2}/\d{2}/\d{4})\s+'              # value date
            r'(\d{2}/\d{2}/\d{4})\s+'              # transaction date
            r'(.+?)\s+'                             # description (greedy but will be trimmed)
            r'([\d,]+\.\d{2})\s+'                  # withdrawal amount
            r'([\d,]+\.\d{2})\s+'                  # deposit amount
            r'([\d,]+\.\d{2})\s*$'                 # balance
        )

        # Simpler pattern: date + description only (amounts on a separate or no-amount continuation)
        date_start_re = re.compile(
            r'(?:\d+\s+)?'
            r'(\d{2}/\d{2}/\d{4})\s+'
            r'(\d{2}/\d{2}/\d{4})\s+'
            r'(.+)'
        )

        # Amount-only tail pattern: matches "withdrawal deposit balance" at end of a line
        amounts_tail_re = re.compile(
            r'([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$'
        )

        i = 0
        while i < len(all_lines):
            line = all_lines[i].strip()
            if not line:
                i += 1
                continue

            # Try full single-line match
            m = txn_re.match(line)
            if m:
                value_date_str, txn_date_str, description, withdrawal_str, deposit_str, balance_str = m.groups()
                date_str = txn_date_str if txn_date_str else value_date_str
                description = description.strip()

                debit = self._clean_amount(withdrawal_str)
                credit = self._clean_amount(deposit_str)
                balance = self._clean_amount(balance_str)

                transaction_date = self._parse_date(date_str, ['%d/%m/%Y'])
                if transaction_date and (debit > 0 or credit > 0):
                    transactions.append({
                        'transaction_date': transaction_date,
                        'description': description,
                        'amount': debit if debit > 0 else credit,
                        'transaction_type': self._detect_transaction_type(description, debit, credit),
                        'payment_method': self._detect_payment_method(description),
                        'merchant_name': self._extract_merchant_name(description),
                        'reference_number': None,
                        'balance_after': balance
                    })
                i += 1
                continue

            # Try multi-line: date line with partial description, continuation lines, then amounts
            dm = date_start_re.match(line)
            if dm:
                value_date_str, txn_date_str, desc_part = dm.groups()
                desc_parts = [desc_part.strip()]

                # Look ahead for continuation lines (no date prefix, no amount-only pattern that ends the block)
                j = i + 1
                withdrawal_str = deposit_str = balance_str = None
                while j < len(all_lines):
                    next_line = all_lines[j].strip()
                    if not next_line:
                        j += 1
                        continue

                    # Check if next line starts a new transaction
                    if date_start_re.match(next_line):
                        break

                    # Check if the current accumulated desc + this line ends with amounts
                    combined = ' '.join(desc_parts) + ' ' + next_line
                    am = amounts_tail_re.search(combined)
                    if am:
                        # Extract description part before amounts
                        desc_before_amounts = combined[:am.start()].strip()
                        if desc_before_amounts:
                            desc_parts = [desc_before_amounts]
                        withdrawal_str, deposit_str, balance_str = am.groups()
                        j += 1
                        break
                    else:
                        # Check if this line itself is just amounts
                        am2 = amounts_tail_re.match(next_line)
                        if am2:
                            withdrawal_str, deposit_str, balance_str = am2.groups()
                            j += 1
                            break
                        # It's a continuation of the description
                        desc_parts.append(next_line)
                    j += 1

                if withdrawal_str and deposit_str and balance_str:
                    date_str = txn_date_str if txn_date_str else value_date_str
                    description = ' '.join(desc_parts).strip()
                    debit = self._clean_amount(withdrawal_str)
                    credit = self._clean_amount(deposit_str)
                    balance = self._clean_amount(balance_str)

                    transaction_date = self._parse_date(date_str, ['%d/%m/%Y'])
                    if transaction_date and (debit > 0 or credit > 0):
                        transactions.append({
                            'transaction_date': transaction_date,
                            'description': description,
                            'amount': debit if debit > 0 else credit,
                            'transaction_type': self._detect_transaction_type(description, debit, credit),
                            'payment_method': self._detect_payment_method(description),
                            'merchant_name': self._extract_merchant_name(description),
                            'reference_number': None,
                            'balance_after': balance
                        })
                    i = j
                    continue

            i += 1

        return transactions
    
    def _parse_excel(self) -> List[Dict[str, Any]]:
        """Parse ICICI Bank Excel statement"""
        transactions = []
        
        # Auto-detect engine based on file extension
        engine = 'xlrd' if self.file_path.lower().endswith('.xls') else 'openpyxl'
        
        # Read the entire file first to find the header row
        df_raw = pd.read_excel(self.file_path, engine=engine, header=None)
        
        # Find the header row (contains multiple transaction-related columns)
        # Look for row with "S No." or combination of "Transaction Date" AND "Balance"
        header_row = None
        for idx, row in df_raw.iterrows():
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            # Check for ICICI statement header pattern
            if ('S No.' in row_str and 'Transaction Date' in row_str) or \
               ('Value Date' in row_str and 'Balance(INR)' in row_str) or \
               ('Transaction Date' in row_str and 'Withdrawal Amount' in row_str):
                header_row = idx
                break
        
        if header_row is None:
            # Fallback: try standard column names
            df = pd.read_excel(self.file_path, engine=engine)
        else:
            # Read again with the correct header row
            df = pd.read_excel(self.file_path, engine=engine, header=header_row)
        
        # Normalize column names: strip whitespace, collapse spaces, remove spaces inside parens
        def _norm_col(c):
            c = re.sub(r'\s+', ' ', str(c).strip())
            c = re.sub(r'\(\s+', '(', c)
            c = re.sub(r'\s+\)', ')', c)
            return c
        df.columns = [_norm_col(c) for c in df.columns]

        # Map possible column names to standard names
        column_mapping = {
            'Transaction Date': 'date',
            'Value Date': 'value_date',
            'Transaction Remarks': 'description',
            'Cheque Number': 'cheque_no',
            'Withdrawal Amount(INR)': 'debit',
            'Withdrawal Amount (INR)': 'debit',
            'Deposit Amount(INR)': 'credit',
            'Deposit Amount (INR)': 'credit',
            'Balance(INR)': 'balance',
            'Balance (INR)': 'balance',
            # Alternative names
            'Date': 'date',
            'Description': 'description',
            'Cheque No': 'cheque_no',
            'Debit': 'debit',
            'Credit': 'credit',
            'Balance': 'balance'
        }

        # Rename columns
        df = df.rename(columns=column_mapping)
        
        # Parse transactions
        for _, row in df.iterrows():
            try:
                # Get date (prefer Transaction Date over Value Date)
                date_str = str(row.get('date', row.get('value_date', '')))
                if not date_str or date_str == 'nan':
                    continue
                
                description = str(row.get('description', ''))
                if not description or description == 'nan':
                    continue
                
                cheque_no = str(row.get('cheque_no', ''))
                debit = self._clean_amount(str(row.get('debit', 0)))
                credit = self._clean_amount(str(row.get('credit', 0)))
                balance = self._clean_amount(str(row.get('balance', 0)))
                
                transaction_date = self._parse_date(date_str, ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'])
                
                if transaction_date:
                    transactions.append({
                        'transaction_date': transaction_date,
                        'description': description.strip(),
                        'amount': debit if debit > 0 else credit,
                        'transaction_type': self._detect_transaction_type(description, debit, credit),
                        'payment_method': self._detect_payment_method(description),
                        'merchant_name': self._extract_merchant_name(description),
                        'reference_number': cheque_no if cheque_no and cheque_no != 'nan' else None,
                        'balance_after': balance
                    })
            except Exception as e:
                continue
        
        return transactions


class ICICICreditCardParser(BankStatementParser):
    """Parser for ICICI Credit Card statements (Excel and PDF format)"""

    def parse(self) -> List[Dict[str, Any]]:
        """Parse ICICI Credit Card statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        elif self.file_path.lower().endswith(('.xlsx', '.xls')):
            return self._parse_excel()
        else:
            raise ValueError("Unsupported file format for ICICI Credit Card. Only PDF and Excel files are supported.")
    
    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse ICICI Credit Card PDF statement using pdfplumber.

        PDF table layout:
          Date | SerNo. | Transaction Details | Reward Points | Intl.# amount | Amount (in₹)

        Amount column contains the value with optional "CR" suffix for credits.
        Card number header lines like "4315XXXXXXXX3003" separate card sections.
        """
        import pdfplumber

        # Try text-based parsing first — more reliable for ICICI CC PDFs
        # where table extraction only picks up highlighted/bordered rows
        transactions = self._parse_pdf_text(pdfplumber)
        if transactions:
            return transactions
        return self._parse_pdf_tables(pdfplumber)

    def _parse_pdf_tables(self, pdfplumber) -> List[Dict[str, Any]]:
        """Try pdfplumber table extraction for ICICI CC PDF."""
        transactions = []
        date_formats = ['%d/%m/%Y', '%d-%m-%Y']

        with pdfplumber.open(self.file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    for row in table:
                        if not row or len(row) < 4:
                            continue
                        # Find the date cell — could be at index 0
                        date_str = (row[0] or '').strip()
                        if not re.match(r'^\d{2}/\d{2}/\d{4}$', date_str):
                            continue

                        transaction_date = self._parse_date(date_str, date_formats)
                        if not transaction_date:
                            continue

                        # Determine column layout:
                        # 6 cols: Date | SerNo | Details | Reward | Intl | Amount
                        # 5 cols: Date | SerNo | Details | Reward | Amount
                        if len(row) >= 6:
                            serial_no = (row[1] or '').strip()
                            description = (row[2] or '').strip()
                            amount_str = (row[-1] or '').strip()
                        elif len(row) >= 5:
                            serial_no = (row[1] or '').strip()
                            description = (row[2] or '').strip()
                            amount_str = (row[-1] or '').strip()
                        else:
                            serial_no = ''
                            description = (row[1] or '').strip()
                            amount_str = (row[-1] or '').strip()

                        if not description or not amount_str:
                            continue

                        # Clean multi-line description
                        description = re.sub(r'\s+', ' ', description).strip()

                        # Parse amount — "27,944.96 CR" or "952.00"
                        is_credit = 'CR' in amount_str.upper()
                        amount_clean = re.sub(r'[CRcr\s]', '', amount_str)
                        amount = self._clean_amount(amount_clean)

                        if amount == 0:
                            continue

                        transaction_type = ExpenseType.CREDIT if is_credit else ExpenseType.DEBIT

                        transactions.append({
                            'transaction_date': transaction_date,
                            'description': description,
                            'amount': amount,
                            'transaction_type': transaction_type,
                            'payment_method': PaymentMethod.CREDIT_CARD,
                            'merchant_name': self._extract_merchant_name(description),
                            'reference_number': serial_no if serial_no and serial_no.isdigit() else None,
                            'balance_after': None
                        })

        return transactions

    def _parse_pdf_text(self, pdfplumber) -> List[Dict[str, Any]]:
        """Fallback text-based parsing for ICICI CC PDF.

        Text lines look like:
          31/10/2025 12246982564 BBPS Payment received 0 27,944.96 CR
          30/10/2025 12248546676 AMAZON PAY IN GROCERY BANGALORE IN 47 952.00
        Or multi-line:
          01/11/2025 12259305108 AMAZON PAY IN E COMMERC BANGALORE 42 843.00
          IN
        Or with chart label noise (SPENDS OVERVIEW pie chart bleeds into text):
          26% 03/11/2025 12273883785 BHARTI AIRTEL LTD GURGAON IN 0 2,358.82
          Travel-26% Apparel/Grocery-6% 04/11/2025 12278569234 AMAZON PAY ... 627.70 CR
        """
        transactions = []
        date_formats = ['%d/%m/%Y', '%d-%m-%Y']

        # Match: DD/MM/YYYY <serial> <description> <reward_pts> [intl_amt] <amount> [CR]
        # Uses re.search (not match) because chart labels may prefix the date
        txn_re = re.compile(
            r'(\d{2}/\d{2}/\d{4})\s+(\d{8,15})\s+(.+?)\s+(-?\d+)\s+'
            r'(?:([\d,]+\.\d{2})\s+)?'  # optional intl amount
            r'([\d,]+\.\d{2})\s*(CR)?\s*$',
            re.IGNORECASE
        )
        # Skip patterns
        skip_re = re.compile(
            r'(Date\s+SerNo|CREDIT CARD|STATEMENT|Page \d|Invoice|Credit Limit|'
            r'Previous Balance|EARNINGS|SPENDS|IMPORTANT|GREAT OFFERS|'
            r'^\d{4}[X]+\d+$|^#|International Spends)',
            re.IGNORECASE
        )
        # Pattern to strip chart label noise that prefixes transaction lines
        chart_noise_re = re.compile(
            r'^(?:[\d.]+%\s*)*(?:(?:Travel|Others|Apparel|Grocery|Shopping|Food|Fuel|Entertainment)'
            r'[-/\w]*[-]?\d*%?\s*)*',
            re.IGNORECASE
        )

        with pdfplumber.open(self.file_path) as pdf:
            all_lines: List[str] = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_lines.extend(text.split('\n'))

        i = 0
        while i < len(all_lines):
            line = all_lines[i].strip()

            if not line or skip_re.search(line):
                i += 1
                continue

            # Try matching directly first, then with chart noise stripped
            match = txn_re.search(line)
            if not match:
                i += 1
                continue

            date_str, serial_no, description, reward_pts, intl_amt, amount_str, cr_flag = match.groups()

            # Collect continuation lines (e.g., "IN" on next line)
            k = i + 1
            while k < len(all_lines):
                nxt = all_lines[k].strip()
                # Strip chart noise from continuation lines too (e.g., "Others-69% IN")
                nxt_clean = chart_noise_re.sub('', nxt).strip()
                # Continuation is a short non-date, non-numeric line
                if (not nxt_clean or skip_re.search(nxt)
                        or re.search(r'\d{2}/\d{2}/\d{4}', nxt_clean)
                        or re.match(r'^[\d,]+\.\d{2}', nxt_clean)):
                    break
                # Short continuation (like "IN" or location suffix)
                if len(nxt_clean) <= 40:
                    description += ' ' + nxt_clean
                    k += 1
                else:
                    break

            transaction_date = self._parse_date(date_str, date_formats)
            if not transaction_date:
                i = k
                continue

            is_credit = cr_flag is not None
            amount = self._clean_amount(amount_str)

            if amount == 0:
                i = k
                continue

            description = re.sub(r'\s+', ' ', description).strip()
            transaction_type = ExpenseType.CREDIT if is_credit else ExpenseType.DEBIT

            transactions.append({
                'transaction_date': transaction_date,
                'description': description,
                'amount': amount,
                'transaction_type': transaction_type,
                'payment_method': PaymentMethod.CREDIT_CARD,
                'merchant_name': self._extract_merchant_name(description),
                'reference_number': serial_no,
                'balance_after': None
            })

            i = k

        return transactions

    def _parse_excel(self) -> List[Dict[str, Any]]:
        """Parse ICICI Credit Card Excel statement"""
        transactions = []

        # Try openpyxl first (for .xlsx), then xlrd (for .xls)
        # Some .xls files are actually .xlsx format
        try:
            df = pd.read_excel(self.file_path, sheet_name='CCLastStatement', engine='openpyxl', header=None)
        except Exception as e:
            logger.debug(f"Excel engine fallback: openpyxl failed, trying xlrd: {e}")
            try:
                df = pd.read_excel(self.file_path, sheet_name='CCLastStatement', engine='xlrd', header=None)
            except Exception as e:
                logger.debug(f"Excel engine fallback: xlrd failed, using auto-detect: {e}")
                # Fallback: let pandas auto-detect
                df = pd.read_excel(self.file_path, sheet_name='CCLastStatement', header=None)
        
        # Find the transaction data start row
        # Look for "Transaction Date" header (row 15 in the sample)
        transaction_start_row = None
        for idx, row in df.iterrows():
            if pd.notna(row.iloc[1]) and 'Transaction Date' in str(row.iloc[1]):
                transaction_start_row = idx + 2  # Data starts 2 rows after header
                break
        
        if transaction_start_row is None:
            return transactions
        
        # Parse transactions
        # Format: Column 1 = Date, Column 5 = Description, Column 9 = Amount, Column 13 = Reference
        for idx in range(transaction_start_row, len(df)):
            try:
                row = df.iloc[idx]
                
                # Check if we've reached the end of transactions (empty date)
                if pd.isna(row.iloc[1]):
                    break
                
                # Extract transaction data
                date_str = str(row.iloc[1]).strip()
                description = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
                amount_str = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                reference = str(row.iloc[13]).strip() if pd.notna(row.iloc[13]) else None
                
                # Skip if no description or amount
                if not description or not amount_str or description == 'nan' or amount_str == 'nan':
                    continue
                
                # Parse date
                transaction_date = self._parse_date(date_str, ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'])
                if not transaction_date:
                    continue
                
                # Parse amount and determine transaction type
                # Format: "833 Dr." or "2 Cr."
                is_credit = 'Cr.' in amount_str or 'Cr' in amount_str
                is_debit = 'Dr.' in amount_str or 'Dr' in amount_str
                
                # Clean amount (remove "Dr.", "Cr.", spaces)
                amount_clean = amount_str.replace('Dr.', '').replace('Cr.', '').replace('Dr', '').replace('Cr', '').strip()
                amount = self._clean_amount(amount_clean)
                
                if amount == 0:
                    continue
                
                # Determine transaction type
                if is_credit:
                    transaction_type = ExpenseType.CREDIT
                    debit = 0.0
                    credit = amount
                else:
                    transaction_type = ExpenseType.DEBIT
                    debit = amount
                    credit = 0.0
                
                # For credit cards, payment method is always CREDIT_CARD
                payment_method = PaymentMethod.CREDIT_CARD
                
                transactions.append({
                    'transaction_date': transaction_date,
                    'description': description,
                    'amount': amount,
                    'transaction_type': transaction_type,
                    'payment_method': payment_method,
                    'merchant_name': self._extract_merchant_name(description),
                    'reference_number': reference if reference and reference != 'nan' else None,
                    'balance_after': None  # Credit card statements don't have running balance
                })
                
            except Exception as e:
                # Skip problematic rows
                continue
        
        return transactions


class ScapiaCreditCardParser(BankStatementParser):
    """Parser for Scapia Credit Card statements (PDF format)"""
    
    def parse(self) -> List[Dict[str, Any]]:
        """Parse Scapia Credit Card statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        else:
            raise ValueError("Unsupported file format for Scapia Credit Card. Only PDF files are supported.")
    
    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse Scapia Credit Card PDF statement using pdfplumber with PyPDF2 fallback.

        Some Scapia statement layouts cause pdfplumber to garble text due to
        overlapping columns, while PyPDF2 extracts cleanly. Try pdfplumber first,
        fall back to PyPDF2 if no transactions are found.
        """
        import pdfplumber

        # Try pdfplumber first
        all_lines = []
        try:
            with pdfplumber.open(self.file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        all_lines.extend(text.split('\n'))
        except Exception as e:
            logger.error(f"Scapia CC PDF pdfplumber extraction failed: {e}")

        transactions = self._extract_transactions(all_lines)
        if transactions:
            return transactions

        # Fallback: PyPDF2 handles some Scapia layouts better
        all_lines = []
        try:
            reader = PyPDF2.PdfReader(self.file_path)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    all_lines.extend(text.split('\n'))
        except Exception as e:
            logger.error(f"Scapia CC PDF PyPDF2 extraction failed: {e}")
            return []

        return self._extract_transactions(all_lines)

    def _extract_transactions(self, all_lines: list) -> List[Dict[str, Any]]:
        """Extract transactions from text lines using Scapia format patterns."""
        transactions = []

        # Scapia format variants (pdfplumber output):
        # "24 Sep 2025 · 08:31 Bistro ₹158.00"
        # "25 Sep 2025 · 17:54 Santosini Dash ₹690.00 35"
        # "30 Sep 2025 · 18:16 Bill payment Payment + ₹15,071.90"
        # PyPDF2 format (no spaces):
        # "24-11-2025 ·19:38ApolloPharmacy ₹270.70"
        # "28-11-2025 ·11:44Billpayment Payment +₹16,969.83"

        # Pattern 1: pdfplumber — "DD Mon YYYY · HH:MM Description [Payment] [+] ₹Amount [RewardPts]"
        pdfplumber_re = re.compile(
            r'(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})\s*[·:]\s*(\d{2}:\d{2})\s+'  # date · time
            r'(.+?)\s+'                                                       # description
            r'(\+\s*)?[₹]\s*([\d,]+\.\d{2})\s*'                              # optional + sign, ₹amount
            r'(?:[\d,]+)?\s*$'                                                # optional reward points
        )

        # Pattern 2: PyPDF2 fallback — "DD-MM-YYYY ·HH:MMDescription ₹Amount"
        pypdf2_re = re.compile(
            r'(\d{2}-\d{2}-\d{4})\s*·\s*(\d{2}:\d{2})(.+?)\s+'  # date·timeDescription
            r'(\+\s*)?[₹]\s*([\d,]+\.?\d*)\s*'                    # optional +, ₹amount
            r'(?:\d+)?\s*$'                                        # optional reward points
        )

        for line in all_lines:
            line = line.strip()
            if not line:
                continue

            # Try pdfplumber format first
            m = pdfplumber_re.match(line)
            if m:
                date_str, time_str, description, plus_sign, amount_str = m.groups()
                transaction_date = self._parse_date(date_str, ['%d %b %Y'])
            else:
                # Try PyPDF2 format
                m = pypdf2_re.match(line)
                if m:
                    date_str, time_str, description, plus_sign, amount_str = m.groups()
                    transaction_date = self._parse_date(date_str, ['%d-%m-%Y'])
                else:
                    continue

            if not transaction_date:
                continue

            description = description.strip()
            if not description:
                continue

            amount = self._clean_amount(amount_str)
            if amount == 0:
                continue

            # Credit detection: "+" prefix or "Payment" in description
            is_credit = bool(plus_sign) or 'Payment' in description

            if is_credit:
                transaction_type = ExpenseType.CREDIT
            else:
                transaction_type = ExpenseType.DEBIT

            # Clean "Payment" suffix from description for bill payments
            desc_clean = re.sub(r'\s+Payment\s*$', '', description).strip()

            transactions.append({
                'transaction_date': transaction_date,
                'description': desc_clean,
                'amount': amount,
                'transaction_type': transaction_type,
                'payment_method': PaymentMethod.CREDIT_CARD,
                'merchant_name': self._extract_merchant_name(desc_clean),
                'reference_number': None,
                'balance_after': None
            })

        logger.info(f"Scapia CC PDF: found {len(transactions)} transactions")
        return transactions


class IDFCFirstCreditCardParser(BankStatementParser):
    """Parser for IDFC First Bank Credit Card statements (PDF format)"""
    
    def parse(self) -> List[Dict[str, Any]]:
        """Parse IDFC First Bank Credit Card statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        else:
            raise ValueError("Unsupported file format for IDFC First Bank Credit Card. Only PDF files are supported.")
    
    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse IDFC First Bank Credit Card PDF statement using pdfplumber.

        Handles multi-line merchant descriptions where pdfplumber splits long
        merchant names across lines, e.g.:
          INNOVATIVE RETAIL CONC,
          09 Feb 26 808.00 DR
          BENGALURU
        """
        import pdfplumber

        transactions = []

        # Single-line: DD Mon YY <description> [Convert] <amount> DR/CR
        single_line_re = re.compile(
            r'(\d{2}\s+[A-Za-z]{3}\s+\d{2})\s+(.+?)\s+(Convert\s+)?[\s\u00a0]*([\d,]+\.\d{2})\s+(DR|CR)'
        )
        # Multi-line: DD Mon YY [Convert] <amount> DR/CR  (date + amount, NO description)
        date_amount_only_re = re.compile(
            r'(\d{2}\s+[A-Za-z]{3}\s+\d{2})\s+(Convert\s+)?([\d,]+\.\d{2})\s+(DR|CR)\s*$'
        )
        # Skip patterns — header/section lines that look like transactions
        skip_re = re.compile(
            r'(Card Number|Purchases|Payments & Other|Transaction|Amount|Eligibility|'
            r'YOUR TRANSACTIONS|YOUR CARD|Credit Card Statement|Statement Date|'
            r'Relationship No|CKYC|^\d{2}/[A-Za-z]{3}/\d{4})'
        )
        # Date prefix pattern — to detect if a line starts with a transaction date
        date_prefix_re = re.compile(r'^\d{2}\s+[A-Za-z]{3}\s+\d{2}')

        with pdfplumber.open(self.file_path) as pdf:
            all_lines: List[str] = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_lines.extend(text.split('\n'))

            i = 0
            while i < len(all_lines):
                line = all_lines[i].strip()

                # Skip headers and non-transaction lines
                if not line or skip_re.search(line):
                    i += 1
                    continue

                # Try single-line match first (most common case)
                match = single_line_re.match(line)
                if match:
                    date_str, description, _, amount_str, dr_cr = match.groups()
                    # "Convert" alone is not a real description — treat as multi-line
                    if description.strip().lower() != 'convert':
                        txn = self._build_transaction(date_str, description, amount_str, dr_cr)
                        if txn:
                            transactions.append(txn)
                        i += 1
                        continue

                # Try multi-line: current line has date + amount but no description
                # Also matches "DD Mon YY Convert AMOUNT DR/CR" (description was on prev line)
                ml_match = date_amount_only_re.match(line)
                if not ml_match:
                    # Re-check with Convert pattern for lines like "13 Feb 26 Convert 1,568.29 DR"
                    ml_match = re.match(
                        r'(\d{2}\s+[A-Za-z]{3}\s+\d{2})\s+Convert\s+([\d,]+\.\d{2})\s+(DR|CR)\s*$',
                        line
                    )
                    if ml_match:
                        date_str, amount_str, dr_cr = ml_match.groups()
                    else:
                        i += 1
                        continue
                else:
                    date_str, _, amount_str, dr_cr = ml_match.groups()

                # Look backward exactly one line for the description prefix
                desc_before = ''
                if i > 0:
                    prev = all_lines[i - 1].strip()
                    if (prev and not skip_re.search(prev)
                            and not date_prefix_re.match(prev)
                            and not re.search(r'\d\s+(DR|CR)\s*$', prev)):
                        desc_before = prev

                # Look forward exactly one line for location continuation
                desc_after = ''
                k = i + 1
                if k < len(all_lines):
                    nxt = all_lines[k].strip()
                    if (nxt and not skip_re.search(nxt)
                            and not date_prefix_re.match(nxt)
                            and not re.search(r'\d\s+(DR|CR)\s*$', nxt)):
                        desc_after = nxt
                        k += 1

                description = (desc_before + ' ' + desc_after).strip() if desc_before else desc_after
                txn = self._build_transaction(date_str, description, amount_str, dr_cr)
                if txn:
                    transactions.append(txn)
                i = k
                continue

        return transactions

    def _build_transaction(self, date_str: str, description: str, amount_str: str, dr_cr: str) -> Optional[Dict[str, Any]]:
        """Build a transaction dict from parsed components."""
        transaction_date = self._parse_date(date_str, ['%d %b %y'])
        if not transaction_date:
            return None

        description = description.strip()
        if not description:
            return None

        amount = self._clean_amount(amount_str)
        if amount == 0:
            return None

        is_credit = dr_cr == 'CR'
        transaction_type = ExpenseType.CREDIT if is_credit else ExpenseType.DEBIT

        return {
            'transaction_date': transaction_date,
            'description': description,
            'amount': amount,
            'transaction_type': transaction_type,
            'payment_method': PaymentMethod.CREDIT_CARD,
            'merchant_name': self._extract_merchant_name(description),
            'reference_number': None,
            'balance_after': None
        }


class HDFCBankParser(BankStatementParser):
    """Parser for HDFC Bank statements (PDF and Excel)"""
    
    def parse(self) -> List[Dict[str, Any]]:
        """Parse HDFC Bank statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        elif self.file_path.lower().endswith(('.xlsx', '.xls')):
            return self._parse_excel()
        else:
            raise ValueError("Unsupported file format for HDFC Bank")
    
    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse HDFC Bank PDF statement"""
        transactions = []
        
        with open(self.file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page in pdf_reader.pages:
                text = page.extract_text()
                lines = text.split('\n')
                
                for line in lines:
                    # HDFC format: Date | Narration | Chq/Ref No | Value Date | Debit | Credit | Balance
                    match = re.match(
                        r'(\d{2}/\d{2}/\d{2})\s+(.+?)\s+(\w+)?\s+(\d{2}/\d{2}/\d{2})\s+([\d,]+\.\d{2})?\s+([\d,]+\.\d{2})?\s+([\d,]+\.\d{2})',
                        line
                    )
                    
                    if match:
                        date_str, narration, ref_no, value_date_str, debit_str, credit_str, balance_str = match.groups()
                        
                        debit = self._clean_amount(debit_str) if debit_str else 0.0
                        credit = self._clean_amount(credit_str) if credit_str else 0.0
                        balance = self._clean_amount(balance_str) if balance_str else 0.0
                        
                        transaction_date = self._parse_date(date_str, ['%d/%m/%y'])
                        
                        if transaction_date:
                            transactions.append({
                                'transaction_date': transaction_date,
                                'description': narration.strip(),
                                'amount': debit if debit > 0 else credit,
                                'transaction_type': self._detect_transaction_type(narration, debit, credit),
                                'payment_method': self._detect_payment_method(narration),
                                'merchant_name': self._extract_merchant_name(narration),
                                'reference_number': ref_no if ref_no else None,
                                'balance_after': balance
                            })
        
        return transactions
    
    def _parse_excel(self) -> List[Dict[str, Any]]:
        """Parse HDFC Bank Excel statement"""
        transactions = []
        
        # Auto-detect engine based on file extension
        engine = 'xlrd' if self.file_path.lower().endswith('.xls') else 'openpyxl'
        
        # Read the entire file first to find the header row
        df_raw = pd.read_excel(self.file_path, engine=engine, header=None)
        
        # Find the header row (contains "Date", "Narration", etc.)
        header_row = None
        for idx, row in df_raw.iterrows():
            row_str = ' '.join([str(x) for x in row if pd.notna(x)])
            # Look for HDFC statement header pattern
            if 'Date' in row_str and 'Narration' in row_str and 'Closing Balance' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            # Fallback: try standard column names
            df = pd.read_excel(self.file_path, engine=engine)
        else:
            # Read again with the correct header row, skip the separator row
            df = pd.read_excel(self.file_path, engine=engine, header=header_row)
        
        # Map possible column names to standard names
        column_mapping = {
            'Date': 'date',
            'Narration': 'description',
            'Chq./Ref.No.': 'ref_no',
            'Chq/Ref No': 'ref_no',
            'Value Dt': 'value_date',
            'Withdrawal Amt.': 'debit',
            'Deposit Amt.': 'credit',
            'Closing Balance': 'balance',
        }
        
        # Rename columns
        df = df.rename(columns=column_mapping)
        
        # HDFC Excel columns: Date, Narration, Chq/Ref No, Value Dt, Withdrawal Amt, Deposit Amt, Closing Balance
        for _, row in df.iterrows():
            try:
                # Get date
                date_str = str(row.get('date', ''))
                if not date_str or date_str == 'nan' or date_str == '********':
                    continue
                
                # Get description
                description = str(row.get('description', ''))
                if not description or description == 'nan' or '**********' in description:
                    continue
                
                # Skip summary rows
                if 'Total' in description or 'End of' in description or 'Generated' in description:
                    continue
                
                ref_no = str(row.get('ref_no', ''))
                debit = self._clean_amount(str(row.get('debit', 0)))
                credit = self._clean_amount(str(row.get('credit', 0)))
                balance = self._clean_amount(str(row.get('balance', 0)))
                
                # Parse date - HDFC uses format like "27/01/26" or "DD/MM/YY"
                transaction_date = self._parse_date(date_str, ['%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'])
                
                if transaction_date and description:
                    transactions.append({
                        'transaction_date': transaction_date,
                        'description': description.strip(),
                        'amount': debit if debit > 0 else credit,
                        'transaction_type': self._detect_transaction_type(description, debit, credit),
                        'payment_method': self._detect_payment_method(description),
                        'merchant_name': self._extract_merchant_name(description),
                        'reference_number': ref_no if ref_no and ref_no != 'nan' else None,
                        'balance_after': balance
                    })
            except Exception as e:
                continue
        
        return transactions


class IDFCFirstBankParser(BankStatementParser):
    """Parser for IDFC First Bank statements (PDF and Excel)"""
    
    def parse(self) -> List[Dict[str, Any]]:
        """Parse IDFC First Bank statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        elif self.file_path.lower().endswith(('.xlsx', '.xls')):
            return self._parse_excel()
        else:
            raise ValueError("Unsupported file format for IDFC First Bank")
    
    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse IDFC First Bank PDF statement using pdfplumber.

        PDF table layout:
          Transaction Date | Value Date | Particulars | Cheque No | Debit | Credit | Balance

        Handles multi-line descriptions where long particulars wrap across rows,
        and determines debit/credit by comparing balance with previous balance
        when table extraction is unavailable.
        """
        import pdfplumber

        transactions = self._parse_pdf_tables(pdfplumber)
        if transactions:
            return transactions

        # Fallback: text-based parsing
        return self._parse_pdf_text(pdfplumber)

    def _parse_pdf_tables(self, pdfplumber) -> List[Dict[str, Any]]:
        """Try pdfplumber table extraction for IDFC First Bank savings PDF."""
        transactions = []
        date_formats = ['%d-%b-%Y', '%d-%B-%Y', '%Y-%m-%d']

        with pdfplumber.open(self.file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table:
                        continue
                    for row in table:
                        if not row or len(row) < 7:
                            continue
                        # row: [txn_date, value_date, particulars, cheque_no, debit, credit, balance]
                        txn_date_str = (row[0] or '').strip()
                        if not txn_date_str:
                            continue

                        transaction_date = self._parse_date(txn_date_str, date_formats)
                        if not transaction_date:
                            continue

                        description = (row[2] or '').strip()
                        if not description or description.lower() in ('opening balance', 'closing balance'):
                            continue

                        debit = self._clean_amount(row[4]) if row[4] else 0.0
                        credit = self._clean_amount(row[5]) if row[5] else 0.0
                        balance = self._clean_amount(row[6]) if row[6] else 0.0

                        if debit == 0 and credit == 0:
                            continue

                        # Clean up multi-line description (table cells may contain newlines)
                        description = re.sub(r'\s+', ' ', description).strip()

                        transactions.append({
                            'transaction_date': transaction_date,
                            'description': description,
                            'amount': debit if debit > 0 else credit,
                            'transaction_type': self._detect_transaction_type(description, debit, credit),
                            'payment_method': self._detect_payment_method(description),
                            'merchant_name': self._extract_merchant_name(description),
                            'reference_number': (row[3] or '').strip() or None,
                            'balance_after': balance
                        })

        return transactions

    def _parse_pdf_text(self, pdfplumber) -> List[Dict[str, Any]]:
        """Fallback text-based parsing for IDFC First Bank savings PDF.

        Text lines look like:
          01-Nov-2025 01-Nov-2025 NACH/ZERODHA BROKING 12,500.00 2,34,390.67
          LTD/KPELUT794AUY42                   (continuation of description)

        We get two amounts: transaction amount and running balance. Debit vs credit
        is determined by comparing the balance with the previous balance.
        """
        transactions = []
        date_formats = ['%d-%b-%Y', '%d-%B-%Y']

        # Match: DD-Mon-YYYY DD-Mon-YYYY <description> <amount> <balance>
        txn_re = re.compile(
            r'^(\d{2}-[A-Za-z]{3,9}-\d{4})\s+(\d{2}-[A-Za-z]{3,9}-\d{4})\s+'
            r'(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$'
        )
        # Match opening balance line to seed prev_balance
        opening_re = re.compile(r'Opening Balance\s+([\d,]+\.\d{2})')
        # Date prefix to detect transaction start lines
        date_prefix_re = re.compile(r'^\d{2}-[A-Za-z]{3,9}-\d{4}\s+\d{2}-[A-Za-z]{3,9}-\d{4}')
        # Skip patterns
        skip_re = re.compile(
            r'(STATEMENT OF ACCOUNT|CUSTOMER|ACCOUNT|REGISTERED OFFICE|Page \d|'
            r'Opening Balance|Closing Balance|Total Debit|Total Credit|'
            r'Transaction\s+Date|Value Date|Particulars|Cheque|^-+$)',
            re.IGNORECASE
        )

        prev_balance = 0.0

        with pdfplumber.open(self.file_path) as pdf:
            all_lines: List[str] = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_lines.extend(text.split('\n'))

        # First pass: find opening balance
        for line in all_lines:
            m = opening_re.search(line)
            if m:
                prev_balance = self._clean_amount(m.group(1))
                break

        # Second pass: parse transactions with continuation lines
        i = 0
        while i < len(all_lines):
            line = all_lines[i].strip()

            if not line or skip_re.search(line):
                i += 1
                continue

            match = txn_re.match(line)
            if not match:
                i += 1
                continue

            txn_date_str, _, description, amount_str, balance_str = match.groups()

            # Collect continuation lines (non-date, non-header lines that follow)
            k = i + 1
            while k < len(all_lines):
                nxt = all_lines[k].strip()
                if (not nxt or skip_re.search(nxt) or date_prefix_re.match(nxt)
                        or re.match(r'^[\d,]+\.\d{2}$', nxt)):
                    break
                description += ' ' + nxt
                k += 1

            transaction_date = self._parse_date(txn_date_str, date_formats)
            if not transaction_date:
                i = k
                continue

            amount = self._clean_amount(amount_str)
            balance = self._clean_amount(balance_str)

            if amount == 0:
                i = k
                continue

            # Determine debit/credit by comparing balance with previous balance
            # balance = prev_balance - debit  OR  balance = prev_balance + credit
            expected_after_debit = prev_balance - amount
            expected_after_credit = prev_balance + amount

            debit_diff = abs(expected_after_debit - balance)
            credit_diff = abs(expected_after_credit - balance)

            if debit_diff < credit_diff:
                debit, credit = amount, 0.0
            else:
                debit, credit = 0.0, amount

            description = re.sub(r'\s+', ' ', description).strip()

            transactions.append({
                'transaction_date': transaction_date,
                'description': description,
                'amount': amount,
                'transaction_type': self._detect_transaction_type(description, debit, credit),
                'payment_method': self._detect_payment_method(description),
                'merchant_name': self._extract_merchant_name(description),
                'reference_number': None,
                'balance_after': balance
            })

            prev_balance = balance
            i = k

        return transactions
    
    def _parse_excel(self) -> List[Dict[str, Any]]:
        """Parse IDFC First Bank Excel statement"""
        transactions = []
        
        # Auto-detect engine based on file extension
        engine = 'xlrd' if self.file_path.lower().endswith('.xls') else 'openpyxl'
        
        # Read all sheets to find the one with transactions
        excel_file = pd.ExcelFile(self.file_path, engine=engine)
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine, header=None)
            
            # Look for the header row with transaction columns
            header_row = None
            for idx, row in df.iterrows():
                row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
                if 'Transaction Date' in row_str and 'Particulars' in row_str:
                    header_row = idx
                    break
            
            if header_row is None:
                continue
            
            # Re-read with correct header - skip rows before header
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine=engine, skiprows=header_row+1, header=None)
            
            # Manually set column names based on IDFC format
            # Columns: Transaction Date, Value Date, Particulars, Cheque No., Debit, Credit, Balance
            if len(df.columns) < 7:
                continue
            
            df.columns = ['Transaction Date', 'Value Date', 'Particulars', 'Cheque No.', 'Debit', 'Credit', 'Balance']
            
            # IDFC Excel columns: Transaction Date, Value Date, Particulars, Cheque No., Debit, Credit, Balance
            for _, row in df.iterrows():
                try:
                    # Use integer indices to access columns directly
                    date_str = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
                    description = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ''
                    ref_no = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ''
                    debit = self._clean_amount(str(row.iloc[4])) if pd.notna(row.iloc[4]) else 0
                    credit = self._clean_amount(str(row.iloc[5])) if pd.notna(row.iloc[5]) else 0
                    balance = self._clean_amount(str(row.iloc[6])) if pd.notna(row.iloc[6]) else 0
                    
                    # Skip if no valid date or description
                    if not date_str or date_str == 'nan' or not description or description == 'nan':
                        continue
                    
                    # Skip summary rows
                    if 'Total' in str(description) or 'End of' in str(description) or 'number of' in str(description).lower():
                        continue
                    
                    # Parse date - IDFC uses format like "01-Dec-2025"
                    transaction_date = self._parse_date(date_str, ['%d-%b-%Y', '%d-%B-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'])
                    
                    if transaction_date and description:
                        transactions.append({
                            'transaction_date': transaction_date,
                            'description': description.strip(),
                            'amount': debit if debit > 0 else credit,
                            'transaction_type': self._detect_transaction_type(description, debit, credit),
                            'payment_method': self._detect_payment_method(description),
                            'merchant_name': self._extract_merchant_name(description),
                            'reference_number': ref_no if ref_no and ref_no != 'nan' else None,
                            'balance_after': balance
                        })
                except Exception as e:
                    continue
        
        return transactions


class SBIBankParser(BankStatementParser):
    """Parser for State Bank of India statements (PDF and Excel)"""
    
    def parse(self) -> List[Dict[str, Any]]:
        """Parse SBI Bank statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        elif self.file_path.lower().endswith(('.xlsx', '.xls')):
            return self._parse_excel()
        else:
            raise ValueError("Unsupported file format for SBI")
    
    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse SBI PDF statement"""
        transactions = []
        
        with open(self.file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            for page in pdf_reader.pages:
                text = page.extract_text()
                lines = text.split('\n')
                
                for line in lines:
                    # SBI format: Txn Date | Value Date | Description | Ref No/Cheque No | Debit | Credit | Balance
                    match = re.match(
                        r'(\d{2}\s+\w{3}\s+\d{4})\s+(\d{2}\s+\w{3}\s+\d{4})\s+(.+?)\s+(\w+)?\s+([\d,]+\.\d{2})?\s+([\d,]+\.\d{2})?\s+([\d,]+\.\d{2})',
                        line
                    )
                    
                    if match:
                        txn_date_str, value_date_str, description, ref_no, debit_str, credit_str, balance_str = match.groups()
                        
                        debit = self._clean_amount(debit_str) if debit_str else 0.0
                        credit = self._clean_amount(credit_str) if credit_str else 0.0
                        balance = self._clean_amount(balance_str) if balance_str else 0.0
                        
                        transaction_date = self._parse_date(txn_date_str, ['%d %b %Y'])
                        
                        if transaction_date:
                            transactions.append({
                                'transaction_date': transaction_date,
                                'description': description.strip(),
                                'amount': debit if debit > 0 else credit,
                                'transaction_type': self._detect_transaction_type(description, debit, credit),
                                'payment_method': self._detect_payment_method(description),
                                'merchant_name': self._extract_merchant_name(description),
                                'reference_number': ref_no if ref_no else None,
                                'balance_after': balance
                            })
        
        return transactions
    
    def _parse_excel(self) -> List[Dict[str, Any]]:
        """Parse SBI Excel statement (supports password-protected .xlsx files).

        SBI statement layout (as seen in practice):
          Rows 1-17  : Account header info (name, account number, IFSC, balance, etc.)
          Row 18     : Column headers — Date | Details | Ref No/Cheque No | Debit | Credit | Balance
          Rows 19+   : One transaction per row until an empty date cell or the summary section
        """
        import io
        import openpyxl

        transactions = []

        try:
            # ── 1. Load workbook (decrypt first if password-protected) ──────────
            with open(self.file_path, 'rb') as fh:
                raw = fh.read()

            file_data = io.BytesIO(raw)
            if self.password:
                try:
                    import msoffcrypto
                    office_file = msoffcrypto.OfficeFile(file_data)
                    office_file.load_key(password=self.password)
                    decrypted = io.BytesIO()
                    office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    file_data = decrypted
                except Exception as exc:
                    raise ValueError(f"Failed to decrypt SBI statement (wrong password?): {exc}")

            wb = openpyxl.load_workbook(file_data, data_only=True)
            ws = wb.active

            # ── 2. Locate the header row ─────────────────────────────────────
            HEADER_KEYWORDS = {'date', 'debit', 'credit', 'balance'}
            header_row_idx = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = {str(c).strip().lower() for c in row if c is not None}
                if HEADER_KEYWORDS.issubset(cells):
                    header_row_idx = row_idx
                    break

            if header_row_idx is None:
                raise ValueError(
                    "Could not find the transaction header row in the SBI statement. "
                    "Expected a row containing: Date, Debit, Credit, Balance."
                )

            # ── 3. Parse transaction rows ────────────────────────────────────
            STOP_KEYWORDS = ('statement summary', 'brought forward', 'total debit', 'closing balance')

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx <= header_row_idx:
                    continue

                date_cell = row[0]

                # Detect end-of-transactions sentinel rows
                row_text = ' '.join(str(c).lower() for c in row if c is not None)
                if any(kw in row_text for kw in STOP_KEYWORDS):
                    break

                # Skip rows without a date
                if date_cell is None or str(date_cell).strip() in ('', 'None'):
                    continue

                # ── Date (DD/MM/YYYY string or datetime object from openpyxl) ─
                if hasattr(date_cell, 'strftime'):
                    transaction_date = date_cell
                else:
                    date_str = str(date_cell).strip()
                    transaction_date = self._parse_date(
                        date_str, ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d %b %Y']
                    )
                    if not transaction_date:
                        continue

                # ── Description (clean embedded newlines / tabs) ──────────────
                desc_raw = str(row[1]).strip() if row[1] is not None else ''
                description = ' '.join(desc_raw.replace('\n', ' ').replace('\t', ' ').split())
                if not description or description == 'None':
                    continue

                # ── Reference number ─────────────────────────────────────────
                ref_raw = str(row[2]).strip() if row[2] is not None else ''
                ref_no = ref_raw if ref_raw not in ('', 'None', 'nan') else None

                # ── Debit / Credit ────────────────────────────────────────────
                debit = self._clean_amount(str(row[3])) if row[3] is not None else 0.0
                credit = self._clean_amount(str(row[4])) if row[4] is not None else 0.0

                if debit == 0.0 and credit == 0.0:
                    continue

                # ── Balance (strip trailing CR / DR suffix) ───────────────────
                bal_raw = str(row[5]).strip() if row[5] is not None else '0'
                bal_clean = re.sub(r'[A-Za-z\s]+$', '', bal_raw)  # remove 'CR', 'DR', etc.
                balance = self._clean_amount(bal_clean)

                transactions.append({
                    'transaction_date': transaction_date,
                    'description': description,
                    'amount': debit if debit > 0 else credit,
                    'transaction_type': self._detect_transaction_type(description, debit, credit),
                    'payment_method': self._detect_payment_method(description),
                    'merchant_name': self._extract_merchant_name(description),
                    'reference_number': ref_no,
                    'balance_after': balance,
                })

        except ValueError:
            raise
        except Exception as exc:
            raise ValueError(f"Error parsing SBI Excel statement: {exc}")

        return transactions


class KotakBankParser(BankStatementParser):
    """Parser for Kotak Mahindra Bank statements (PDF)"""

    def parse(self) -> List[Dict[str, Any]]:
        """Parse Kotak Mahindra Bank statement"""
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        else:
            raise ValueError("Unsupported file format for Kotak Mahindra Bank. Only PDF files are supported.")

    def _extract_account_info(self, first_page_text: str) -> Dict[str, str]:
        """Extract account details from the first page header."""
        info: Dict[str, str] = {}

        # Account number — appears as "Account No. 9411675197"
        m = re.search(r'Account\s*No\.?\s*(\d+)', first_page_text)
        if m:
            info['account_number'] = m.group(1)

        # Account holder name — first text line after "Account Statement" date line
        m = re.search(r'\d{2}\s+\w{3}\s+\d{4}\s*-\s*\d{2}\s+\w{3}\s+\d{4}\s*\n(.+?)(?:\s+Account\s*No)', first_page_text)
        if m:
            info['account_holder'] = m.group(1).strip()

        # IFSC code
        m = re.search(r'IFSC\s*Code\s*(\w+)', first_page_text)
        if m:
            info['ifsc_code'] = m.group(1)

        # Branch
        m = re.search(r'Branch\s+(\S.+)', first_page_text)
        if m:
            info['branch'] = m.group(1).strip()

        # Statement period
        m = re.search(r'(\d{2}\s+\w{3}\s+\d{4})\s*-\s*(\d{2}\s+\w{3}\s+\d{4})', first_page_text)
        if m:
            info['period_from'] = m.group(1)
            info['period_to'] = m.group(2)

        return info

    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse Kotak Mahindra Bank PDF statement using pdfplumber for table extraction."""
        import pdfplumber

        transactions: List[Dict[str, Any]] = []
        self.opening_balance: Optional[float] = None

        # Kotak date formats: "02 Jan 2026", "2 Jan 2026", "02-Jan-2026"
        date_formats = ['%d %b %Y', '%d-%b-%Y', '%d/%m/%Y', '%d-%m-%Y']

        with pdfplumber.open(self.file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()

                for table in tables:
                    if not table or len(table) < 2:
                        continue

                    # Identify the transaction table by looking for the header row
                    # Kotak header: ['#', 'Date', 'Description', 'Chq/Ref. No.', 'Withdrawal (Dr.)', 'Deposit (Cr.)', 'Balance']
                    header_row_idx = None
                    for idx, row in enumerate(table):
                        row_text = ' '.join(str(cell or '') for cell in row).lower()
                        if 'date' in row_text and ('withdrawal' in row_text or 'deposit' in row_text) and 'balance' in row_text:
                            header_row_idx = idx
                            break

                    if header_row_idx is None:
                        continue

                    # Process data rows after the header
                    for row in table[header_row_idx + 1:]:
                        try:
                            if len(row) < 7:
                                continue

                            serial_no = str(row[0] or '').strip()
                            date_str = str(row[1] or '').strip()
                            description = str(row[2] or '').strip()
                            ref_no = str(row[3] or '').strip()
                            withdrawal_str = str(row[4] or '').strip()
                            deposit_str = str(row[5] or '').strip()
                            balance_str = str(row[6] or '').strip()

                            # Capture the opening balance, then skip this row
                            if 'opening balance' in description.lower():
                                if balance_str and balance_str != '-':
                                    self.opening_balance = self._clean_amount(balance_str)
                                continue

                            # Skip empty/summary rows
                            if not date_str or date_str == '-':
                                continue

                            # Parse date
                            transaction_date = self._parse_date(date_str, date_formats)
                            if not transaction_date:
                                continue

                            # Parse amounts
                            debit = self._clean_amount(withdrawal_str) if withdrawal_str and withdrawal_str != '-' else 0.0
                            credit = self._clean_amount(deposit_str) if deposit_str and deposit_str != '-' else 0.0
                            balance = self._clean_amount(balance_str) if balance_str and balance_str != '-' else 0.0

                            if debit == 0.0 and credit == 0.0:
                                continue

                            # Clean reference number
                            clean_ref = ref_no if ref_no and ref_no != '-' and ref_no.lower() != 'none' else None

                            transactions.append({
                                'transaction_date': transaction_date,
                                'description': description,
                                'amount': debit if debit > 0 else credit,
                                'transaction_type': self._detect_transaction_type(description, debit, credit),
                                'payment_method': self._detect_payment_method(description),
                                'merchant_name': self._extract_merchant_name(description),
                                'reference_number': clean_ref,
                                'balance_after': balance,
                            })
                        except Exception:
                            continue

        # If pdfplumber table extraction found nothing, fall back to text-based parsing
        if not transactions:
            transactions = self._parse_pdf_text_fallback()

        return transactions

    def _parse_pdf_text_fallback(self) -> List[Dict[str, Any]]:
        """Fallback text-based parsing for Kotak statements when table extraction fails."""
        transactions: List[Dict[str, Any]] = []

        date_formats = ['%d %b %Y', '%d-%b-%Y', '%d/%m/%Y']

        with open(self.file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)

            for page in pdf_reader.pages:
                text = page.extract_text()
                if not text:
                    continue
                lines = text.split('\n')

                for line in lines:
                    # Kotak text line pattern:
                    # <serial> <DD Mon YYYY> <description> <ref_no> <withdrawal> <deposit> <balance>
                    # Example: 1 02 Jan 2026 UPI-MERCHANT-REF123 REF456 500.00 - 13,937.04
                    # Also handle lines without serial number

                    # Pattern: number + date (DD Mon YYYY) + description + amounts at end
                    match = re.match(
                        r'(?:\d+\s+)?'                           # optional serial number
                        r'(\d{1,2}\s+\w{3}\s+\d{4})\s+'         # date: DD Mon YYYY
                        r'(.+?)\s+'                              # description (non-greedy)
                        r'([\d,]+\.\d{2})?\s*'                   # withdrawal (optional)
                        r'([\d,]+\.\d{2})?\s*'                   # deposit (optional)
                        r'([\d,]+\.\d{2})\s*$',                  # balance (required)
                        line.strip()
                    )

                    if not match:
                        continue

                    date_str, description, withdrawal_str, deposit_str, balance_str = match.groups()

                    # Skip opening balance
                    if 'opening balance' in description.lower():
                        continue

                    transaction_date = self._parse_date(date_str, date_formats)
                    if not transaction_date:
                        continue

                    debit = self._clean_amount(withdrawal_str) if withdrawal_str else 0.0
                    credit = self._clean_amount(deposit_str) if deposit_str else 0.0
                    balance = self._clean_amount(balance_str) if balance_str else 0.0

                    if debit == 0.0 and credit == 0.0:
                        continue

                    # Try to extract ref number from description
                    ref_match = re.search(r'(?:Ref\.?\s*No\.?\s*|Chq\.?\s*No\.?\s*)(\S+)', description)
                    ref_no = ref_match.group(1) if ref_match else None

                    transactions.append({
                        'transaction_date': transaction_date,
                        'description': description.strip(),
                        'amount': debit if debit > 0 else credit,
                        'transaction_type': self._detect_transaction_type(description, debit, credit),
                        'payment_method': self._detect_payment_method(description),
                        'merchant_name': self._extract_merchant_name(description),
                        'reference_number': ref_no,
                        'balance_after': balance,
                    })

        return transactions


class AxisBankParser(BankStatementParser):
    """Parser for Axis Bank statements (PDF)"""

    def parse(self) -> List[Dict[str, Any]]:
        if self.file_path.lower().endswith('.pdf'):
            return self._parse_pdf()
        else:
            raise ValueError("Unsupported file format for Axis Bank. Only PDF files are supported.")

    def _extract_account_info(self, first_page_text: str) -> Dict[str, str]:
        """Extract account details from the first page header."""
        info: Dict[str, str] = {}

        # Account holder name — first non-empty line
        lines = first_page_text.split('\n')
        if lines:
            info['account_holder_name'] = lines[0].strip()

        # Account number — "Statement of Axis Account No: 915010034176403"
        m = re.search(r'Account\s*No[:\s]*(\d+)', first_page_text)
        if m:
            info['account_number'] = m.group(1)

        # IFSC Code
        m = re.search(r'IFSC\s*Code[:\s]*(\S+)', first_page_text)
        if m:
            info['ifsc_code'] = m.group(1)

        # Account type from Scheme line — "SB-..." = Savings, "CA-..." = Current
        m = re.search(r'Scheme[:\s]*(\S+)', first_page_text)
        if m:
            scheme = m.group(1).upper()
            if scheme.startswith('SB'):
                info['account_type'] = 'savings'
            elif scheme.startswith('CA'):
                info['account_type'] = 'current'
            else:
                info['account_type'] = 'savings'

        return info

    def _parse_pdf(self) -> List[Dict[str, Any]]:
        """Parse Axis Bank PDF statement using pdfplumber table extraction."""
        import pdfplumber

        transactions: List[Dict[str, Any]] = []
        self.opening_balance: Optional[float] = None
        self.account_info: Dict[str, str] = {}

        date_formats = ['%d-%m-%Y', '%d/%m/%Y', '%d-%b-%Y']

        with pdfplumber.open(self.file_path) as pdf:
            # Extract account info from first page header
            if pdf.pages:
                first_page_text = pdf.pages[0].extract_text() or ''
                self.account_info = self._extract_account_info(first_page_text)

            for page in pdf.pages:
                tables = page.extract_tables()

                for table in tables:
                    if not table or len(table) < 2:
                        continue

                    # Find header row: look for 'Tran Date' and 'Particulars'
                    header_row_idx = None
                    for idx, row in enumerate(table):
                        row_text = ' '.join(str(cell or '') for cell in row).lower()
                        if 'tran date' in row_text and 'particulars' in row_text and 'balance' in row_text:
                            header_row_idx = idx
                            break

                    if header_row_idx is None:
                        continue

                    for row in table[header_row_idx + 1:]:
                        try:
                            if len(row) < 6:
                                continue

                            date_str = str(row[0] or '').strip()
                            chq_no = str(row[1] or '').strip()
                            description = str(row[2] or '').replace('\n', ' ').strip()

                            if not description:
                                continue

                            desc_lower = description.lower()

                            # Capture opening balance
                            if 'opening balance' in desc_lower:
                                bal_str = str(row[5] or '').strip()
                                if bal_str:
                                    self.opening_balance = self._clean_amount(bal_str)
                                continue

                            # Skip summary rows
                            if 'transaction total' in desc_lower or 'closing balance' in desc_lower:
                                continue

                            # Must have a date
                            if not date_str:
                                continue

                            transaction_date = self._parse_date(date_str, date_formats)
                            if not transaction_date:
                                continue

                            debit_str = str(row[3] or '').strip()
                            credit_str = str(row[4] or '').strip()
                            balance_str = str(row[5] or '').strip()

                            debit = self._clean_amount(debit_str) if debit_str else 0.0
                            credit = self._clean_amount(credit_str) if credit_str else 0.0
                            balance = self._clean_amount(balance_str) if balance_str else 0.0

                            if debit == 0.0 and credit == 0.0:
                                continue

                            ref_no = chq_no if chq_no and chq_no != '-' else None

                            transactions.append({
                                'transaction_date': transaction_date,
                                'description': description,
                                'amount': debit if debit > 0 else credit,
                                'transaction_type': self._detect_transaction_type(description, debit, credit),
                                'payment_method': self._detect_payment_method(description),
                                'merchant_name': self._extract_merchant_name(description),
                                'reference_number': ref_no,
                                'balance_after': balance,
                            })
                        except Exception:
                            continue

        return transactions


def get_parser(bank_name: str, file_path: str, password: str = None) -> BankStatementParser:
    """Factory function to get appropriate parser based on bank name"""
    parsers = {
        'ICICI': ICICIBankParser,
        'ICICI_CC': ICICICreditCardParser,
        'SCAPIA_CC': ScapiaCreditCardParser,
        'HDFC': HDFCBankParser,
        'IDFC_FIRST': IDFCFirstBankParser,
        'IDFC_FIRST_CC': IDFCFirstCreditCardParser,
        'SBI': SBIBankParser,
        'KOTAK': KotakBankParser,
        'AXIS': AxisBankParser,
    }

    parser_class = parsers.get(bank_name.upper())
    if not parser_class:
        raise ValueError(f"Unsupported bank: {bank_name}")

    return parser_class(file_path, password=password)

# Made with Bob